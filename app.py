import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from google import genai
from google.genai import types
from io import BytesIO
import os
import time
import io
import fitz  # PyMuPDF
import pytesseract
from PIL import Image
import pdfplumber
from dotenv import load_dotenv
import re
import json
from pathlib import Path
import datetime
import hashlib

# Cargar variables de entorno
load_dotenv()

# ============================================
# CONFIGURACIÓN
# ============================================
st.set_page_config(
    page_title="Transcripción Inteligente de Pólizas",
    page_icon="📋",
    layout="wide"
)

# Colores para Excel (exactamente como en el ejemplo)
COLOR_GRUPO = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
COLOR_TITULO = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
COLOR_COBERTURA = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
COLOR_DEDUCIBLE = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
COLOR_CLAUSULA = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
COLOR_ACLARACION = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
COLOR_VALOR = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# CSS
st.markdown("""
<style>
    .main-header { font-size: 2.5rem; font-weight: bold; color: #1f77b4; text-align: center; }
    .sub-header { font-size: 1.2rem; color: #666; text-align: center; margin-bottom: 2rem; }
    .stProgress > div > div > div > div { background-color: #1f77b4; }
    .success-box { padding: 1rem; background-color: #d4edda; border-left: 5px solid #28a745; margin: 1rem 0; }
    .warning-box { padding: 1rem; background-color: #fff3cd; border-left: 5px solid #ffc107; margin: 1rem 0; }
    .error-box { padding: 1rem; background-color: #f8d7da; border-left: 5px solid #dc3545; margin: 1rem 0; }
</style>
""", unsafe_allow_html=True)

# ============================================
# FUNCIÓN PARA OBTENER TOTAL DE PÁGINAS
# ============================================

def obtener_total_paginas_pdf(file_bytes):
    """Obtiene el número total de páginas de un PDF"""
    try:
        # Intentar con PyMuPDF (más rápido)
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        total = len(doc)
        doc.close()
        return total
    except:
        try:
            # Fallback con pdfplumber
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                return len(pdf.pages)
        except:
            return 0

# ============================================
# FUNCIONES DE EXTRACCIÓN MEJORADAS
# ============================================

def extraer_pdf_texto_completo(file_bytes, pagina_inicio=1, pagina_fin=None):
    """Extrae TODO el texto del PDF preservando el formato y estructura"""
    try:
        texto_completo = []
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            total_paginas = len(pdf.pages)

            inicio = max(1, pagina_inicio) - 1
            fin = pagina_fin if pagina_fin else total_paginas
            fin = min(fin, total_paginas)

            for i in range(inicio, fin):
                pagina = pdf.pages[i]
                texto = pagina.extract_text()
                if texto:
                    # Preservar número de página para referencia
                    texto_completo.append(f"--- PÁGINA {i+1} ---\n{texto}")

        return "\n\n".join(texto_completo) if texto_completo else None
    except Exception as e:
        st.error(f"Error en extracción de texto: {e}")
        return None

def extraer_pdf_ocr_mejorado(file_bytes, pagina_inicio=1, pagina_fin=None, progress_bar=None):
    """OCR optimizado con mejor calidad de imagen"""
    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        total_paginas = len(doc)

        inicio = max(0, pagina_inicio - 1)
        fin = pagina_fin if pagina_fin else total_paginas
        fin = min(fin, total_paginas)

        texto_paginas = []

        for idx, i in enumerate(range(inicio, fin)):
            if progress_bar:
                progress_bar.progress((idx + 1) / (fin - inicio))

            # Mayor DPI para mejor OCR
            pix = doc[i].get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x zoom para mejor calidad
            img = Image.open(io.BytesIO(pix.tobytes("png")))

            # Configuración optimizada de Tesseract
            config = "--oem 3 --psm 6 -l spa"
            texto = pytesseract.image_to_string(img, config=config)
            texto_paginas.append(f"--- PÁGINA {i+1} ---\n{texto}")

        doc.close()
        return "\n\n".join(texto_paginas) if texto_paginas else None
    except Exception as e:
        st.error(f"Error en OCR: {e}")
        return None

# ============================================
# FUNCIÓN DE IA MEJORADA - ESTRUCTURA COMPLETA
# ============================================

def inicializar_cliente():
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        st.error("❌ No se encontró GEMINI_API_KEY en las variables de entorno")
        st.stop()
    return genai.Client(api_key=api_key)

def estructurar_poliza_por_secciones(texto_pdf, nombre_pdf, client, progress_callback=None):
    """
    Estructura la póliza procesando por secciones para mantener contexto
    y evitar truncamiento
    """

    # Dividir en páginas para procesamiento incremental
    paginas = texto_pdf.split("--- PÁGINA ")

    elementos_totales = []
    contexto_acumulado = ""

    for idx, pagina_raw in enumerate(paginas[1:], 1):  # Saltar primera parte vacía
        if not pagina_raw.strip():
            continue

        # Extraer número de página y contenido
        lineas = pagina_raw.split("\n", 1)
        num_pagina = lineas[0].replace("---", "").strip()
        contenido_pagina = lineas[1] if len(lineas) > 1 else ""

        if progress_callback:
            progress_callback(f"Procesando página {num_pagina}...")

        # Acumular contexto de páginas anteriores (últimas 3 páginas)
        contexto = "\n".join(contexto_acumulado.split("\n")[-100:])  # Últimas 100 líneas

        prompt = f"""Eres un experto en análisis de pólizas de seguros. Tu tarea es extraer la información EXACTAMENTE como aparece en el documento, SIN RESUMIR, SIN TRADUCIR, SIN MODIFICAR.

=== CONTEXTO DE PÁGINAS ANTERIORES (PARA CONTINUIDAD) ===
{contexto[-2000:] if contexto else "[Inicio del documento]"}

=== CONTENIDO DE LA PÁGINA ACTUAL {num_pagina} ===
{contenido_pagina[:15000]}  # Limitar para no exceder tokens

=== INSTRUCCIONES CRÍTICAS ===

1. EXTRAE EL TEXTO EXACTAMENTE COMO APARECE, palabra por palabra si es necesario
2. NO resumas, NO parafrasees, NO omitas detalles
3. Identifica cada elemento con estos tipos EXACTOS:
   - "GRUPO": Solo para el encabezado principal "TÉRMINOS Y CONDICIONES"
   - "TITULO": Títulos de sección como "UBICACIÓN DEL RIESGO", "MATERIA DEL SEGURO", "AMPARO 1", etc.
   - "COBERTURA": Descripciones completas de coberturas (copiar TODO el texto)
   - "DEDUCIBLE": Condiciones de deducibles
   - "CLAUSULAS": Cláusulas adicionales completas
   - "ACLARACIONES": Aclaraciones o condiciones especiales
   - "VIGENCIA", "PRIMA", "VALOR ASEGURADO": Información administrativa

4. Para CADA elemento, devuelve un objeto JSON con estas 5 propiedades EXACTAS:
   - "tipo": uno de los valores anteriores
   - "col_b": texto completo de la columna B (descripción)
   - "col_c": texto para columna C (sublímites, usualmente vacío)
   - "col_d": texto para columna D (valores, montos, deducibles)
   - "col_e": texto para columna E (propuesta/condiciones del proponente)

5. Si una cobertura ocupa varias líneas en el PDF, ÚNELAS en "col_b" con espacios
6. Los valores numéricos (USD, porcentajes) van en "col_d"
7. Si ves "OTORGA" o "NO OTORGA", va en "col_e"

=== EJEMPLO DE FORMATO DE SALIDA ===
[
  {{
    "tipo": "TITULO",
    "col_b": "UBICACIÓN DEL RIESGO",
    "col_c": "",
    "col_d": "",
    "col_e": ""
  }},
  {{
    "tipo": "COBERTURA",
    "col_b": "TODOS LOS LOCALES Y/O PREDIOS DONDE EL ASEGURADO DESARROLLE SUS ACTIVIDADES, DE FORMA PERMANENTE O TEMPORAL, SEAN DE SU PROPIEDAD O NO, O SEAN PREDIOS DE TERCEROS BAJO RESPONSABILIDAD, CUSTODIA Y/O CONTROL, ASÍ COMO CUALQUIER OTRO LUGAR DENTRO DEL TERRITORIO NACIONAL, SIN LIMITACIÓN:",
    "col_c": "",
    "col_d": "",
    "col_e": "TODOS LOS LOCALES Y/O PREDIOS DONDE EL ASEGURADO DESARROLLE SUS ACTIVIDADES, DE FORMA PERMANENTE O TEMPORAL, SEAN DE SU PROPIEDAD O NO, O SEAN PREDIOS DE TERCEROS BAJO RESPONSABILIDAD, CUSTODIA Y/O CONTROL, ASÍ COMO CUALQUIER OTRO LUGAR DENTRO DEL TERRITORIO NACIONAL"
  }},
  {{
    "tipo": "COBERTURA",
    "col_b": "EDIFICIOS E INSTALACIONES (AGUA, ELECTRIFICACIÓN, GAS, SEGURIDAD Y REDES SIMILARES, ETC.), CONSTRUCCIONES, OBRAS CIVILES EN GENERAL, INCLUYENDO MEJORAS, FUNDACIONES, GAVIONES, MUROS Y/O MALLAS PERIMETRALES Y DE CONTENCIÓN Y/O CERCOS, REJAS, BARDAS, ENMALLADOS, POSTES, VEREDAS, ACERAS, VÍAS DE ACCESO, INSTALACIONES DE ALMACENAJE, DEPÓSITOS, TANQUES DE AGUA, CAÑERÍAS, TUBERÍAS, DRENAJES, DUCTOS E INSTALACIONES SOBRE Y POR DEBAJO LA TIERRA (DENTRO DE LOS PREDIOS DEL ASEGURADO Y HASTA 50 METROS FUERA) Y CUALQUIER INSTALACIÓN PERMANENTE O TEMPORAL, ELÉCTRICA Y/O MECÁNICA QUE FORMEN PARTE DE LOS EDIFICIOS SU FUNCIONAMIENTO, INCLUYENDO VIDRIOS Y/O CRISTALES Y/O ESPEJOS Y/O VITRALES Y/O CERÁMICAS Y/O LETREROS Y/O GIGANTOGRAFIAS (INCLUYENDO SUS ESTRUCTURAS) Y/O CUALQUIER OTRO",
    "col_c": "",
    "col_d": "1350000",
    "col_e": "EDIFICIOS E INSTALACIONES (AGUA, ELECTRIFICACIÓN, GAS, SEGURIDAD Y REDES SIMILARES), CONSTRUCCIONES, OBRAS CIVILES EN GENERAL, INCLUYENDO FUNDACIONES, GAVIONES, MUROS Y/O MALLAS PERIMETRALES Y DE CONTENCIÓN Y/O CERCOS, REJAS, BARDAS, ENMALLADOS, POSTES, VEREDAS, ACERAS, VÍAS DE ACCESO, INSTALACIONES DE ALMACENAJE, DEPÓSITOS, TANQUES DE AGUA, CAÑERÍAS, TUBERÍAS, DRENAJES, DUCTOS E INSTALACIONES SOBRE Y POR DEBAJO LA TIERRA (DENTRO DE LOS PREDIOS DEL ASEGURADO Y HASTA 20 METROS FUERA) Y CUALQUIER INSTALACIÓN PERMANENTE O TEMPORAL, ELÉCTRICA Y/O MECÁNICA QUE FORMEN PARTE DE LOS EDIFICIOS SU FUNCIONAMIENTO, INCLUYENDO VIDRIOS Y/O CRISTALES Y/O ESPEJOS Y/O VITRALES Y/O CERÁMICAS Y/O LETREROS Y/O GIGANTOGRAFIAS (INCLUYENDO SUS ESTRUCTURAS)"
  }}
]

Devuelve SOLO el array JSON, sin texto adicional antes o después."""

        try:
            response = client.models.generate_content(
                model="models/gemini-3.1-flash-lite-preview",
                contents=prompt,
                config=types.GenerateContentConfig(
                    temperature=0.0,
                    max_output_tokens=8192,
                    top_p=0.1,
                )
            )

            if not response or not response.text:
                continue

            texto_respuesta = response.text.strip()

            # Limpiar JSON
            texto_respuesta = re.sub(r'```json\s*', '', texto_respuesta)
            texto_respuesta = re.sub(r'```\s*', '', texto_respuesta)

            # Extraer array JSON
            inicio = texto_respuesta.find('[')
            fin = texto_respuesta.rfind(']') + 1

            if inicio != -1 and fin > inicio:
                json_str = texto_respuesta[inicio:fin]
                elementos = json.loads(json_str)

                if isinstance(elementos, list):
                    elementos_totales.extend(elementos)
                    contexto_acumulado += "\n" + contenido_pagina

        except Exception as e:
            st.warning(f"Error procesando página {num_pagina}: {e}")
            continue

    return elementos_totales

def validar_completitud(elementos, texto_original):
    """
    Valida que no se hayan perdido secciones importantes
    """
    # Verificar que existan los tipos mínimos
    tipos_encontrados = set(e.get('tipo', '') for e in elementos)
    tipos_requeridos = {'TITULO', 'COBERTURA'}

    faltantes = tipos_requeridos - tipos_encontrados
    if faltantes:
        st.warning(f"⚠️ No se detectaron elementos de tipo: {', '.join(faltantes)}")

    # Verificar longitud promedio de textos
    longitudes = [len(e.get('col_b', '')) for e in elementos if e.get('col_b')]
    if longitudes:
        avg_len = sum(longitudes) / len(longitudes)
        if avg_len < 50:
            st.warning(f"⚠️ Los textos extraídos son muy cortos (promedio: {avg_len:.0f} chars). Posible truncamiento.")

    return len(faltantes) == 0

# ============================================
# CREACIÓN DE EXCEL MEJORADA
# ============================================

def crear_excel_estructurado(datos_por_pdf):
    """Crea Excel con formato exacto al ejemplo, preservando las 5 columnas"""
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for pdf_nombre, elementos in datos_por_pdf.items():
            if not elementos:
                continue

            # Preparar datos en formato de 5 columnas
            filas = []

            # Encabezado del sistema de gestión (como en el ejemplo)
            filas.append({
                'A': '',
                'B': '',
                'C': 'SISTEMA DE GESTIÓN DE CALIDAD',
                'D': '',
                'E': 'COD: OPE-REG-002\nFecha: 03- jun - 13\nVersión: 01'
            })
            filas.append({
                'A': '',
                'B': '',
                'C': 'PÓLIZA PROPUESTA',
                'D': '',
                'E': ''
            })
            filas.append({
                'A': '',
                'B': '',
                'C': '',
                'D': '',
                'E': ''
            })

            # Agregar elementos extraídos
            for elem in elementos:
                filas.append({
                    'A': elem.get('tipo', ''),
                    'B': elem.get('col_b', ''),
                    'C': elem.get('col_c', ''),
                    'D': elem.get('col_d', ''),
                    'E': elem.get('col_e', '')
                })

            df = pd.DataFrame(filas)

            # Limpiar nombre de hoja
            sheet_name = pdf_nombre[:31].replace('.pdf', '').replace('/', '_').replace('\\', '_').replace(':', '-')

            # Guardar sin encabezados
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=0)

            # Aplicar formato
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Bordes
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Formato por tipo de fila
            for row_idx in range(1, len(df) + 1):
                tipo = worksheet.cell(row=row_idx, column=1).value

                # Determinar color según tipo
                if tipo == 'GRUPO':
                    fill = COLOR_GRUPO
                    font = Font(bold=True, size=11)
                elif tipo == 'TITULO' or tipo == 'TÍTULO':
                    fill = COLOR_TITULO
                    font = Font(color="FFFFFF", bold=True, size=11)
                elif tipo == 'COBERTURA':
                    fill = COLOR_COBERTURA
                    font = Font(size=10)
                elif tipo == 'DEDUCIBLE':
                    fill = COLOR_DEDUCIBLE
                    font = Font(size=10)
                elif tipo == 'CLAUSULAS' or tipo == 'CLAUSULA':
                    fill = COLOR_CLAUSULA
                    font = Font(size=10)
                elif tipo == 'ACLARACIONES' or tipo == 'ACLARACION':
                    fill = COLOR_ACLARACION
                    font = Font(size=10)
                else:
                    fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    font = Font(size=10)

                # Aplicar a todas las columnas
                for col in range(1, 6):
                    cell = worksheet.cell(row=row_idx, column=col)
                    cell.fill = fill
                    cell.font = font
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

                # Destacar valores en columna D
                valor_cell = worksheet.cell(row=row_idx, column=4)
                if valor_cell.value and str(valor_cell.value).strip():
                    if any(c.isdigit() for c in str(valor_cell.value)):
                        valor_cell.font = Font(bold=True, color="006100", size=10)
                        valor_cell.fill = COLOR_VALOR

            # Ajustar anchos de columna
            worksheet.column_dimensions['A'].width = 35
            worksheet.column_dimensions['B'].width = 80
            worksheet.column_dimensions['C'].width = 25
            worksheet.column_dimensions['D'].width = 25
            worksheet.column_dimensions['E'].width = 80

            worksheet.freeze_panes = 'A4'

    output.seek(0)
    return output

# ============================================
# INTERFAZ PRINCIPAL
# ============================================

def main():
    st.markdown('<div class="main-header">📋 Comparador Inteligente de Pólizas Antiguas y Nuevas</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-header">Extrae el contenido completo de pólizas PDF al formato Excel exacto</div>', unsafe_allow_html=True)

    # Verificar API Key
    if not os.getenv("GEMINI_API_KEY"):
        st.error("❌ Configura GEMINI_API_KEY en archivo .env")
        st.stop()

    # Sidebar con configuración
    with st.sidebar:
        st.header("⚙️ Configuración Avanzada")

        modo_extraccion = st.radio(
            "Modo de extracción:",
            ["Texto nativo (PDF digital)", "OCR (PDF escaneado)"],
            index=0,
            help="Texto nativo es más rápido y preciso para PDFs digitales"
        )

        st.markdown("---")

        # Opciones de procesamiento
        st.subheader("Opciones de IA")
        validar_resultados = st.checkbox("Validar completitud de extracción", value=True)

        st.info("""
        💡 **Tips:**
        - Para pólizas largas, usa rango de páginas
        - La validación detecta posibles truncamientos
        """)

    # Área de carga
    st.subheader("📁 Cargar Pólizas PDF")

    uploaded_pdfs = st.file_uploader(
        "Selecciona uno o varios archivos PDF de pólizas",
        type=['pdf'],
        accept_multiple_files=True,
        help="Arrastra los archivos o haz clic para seleccionar"
    )

    if uploaded_pdfs:
        st.success(f"✅ {len(uploaded_pdfs)} archivo(s) cargado(s)")

        # Mostrar detalles de archivos
        with st.expander("Ver detalles de archivos"):
            for pdf in uploaded_pdfs:
                size_kb = len(pdf.getvalue()) / 1024
                total_paginas = obtener_total_paginas_pdf(pdf.getvalue())
                st.write(f"📄 {pdf.name} ({size_kb:.1f} KB) - {total_paginas} páginas")

        # Selector de rango de páginas
        st.markdown("---")
        st.subheader("📄 Seleccionar rango de páginas a procesar")

        col1, col2 = st.columns(2)
        with col1:
            pagina_inicio = st.number_input("Página inicial", min_value=1, value=1)
        with col2:
            # Usar el total del primer PDF como valor por defecto
            total_ref = obtener_total_paginas_pdf(uploaded_pdfs[0].getvalue())
            pagina_fin = st.number_input("Página final",
                                        min_value=pagina_inicio,
                                        value=total_ref,  # AHORA ES EL TOTAL DE PÁGINAS
                                        help=f"El primer PDF tiene {total_ref} páginas totales")

        st.caption(f"💡 Se procesarán las páginas {pagina_inicio} a {pagina_fin} de cada PDF")

        # Botón de procesamiento
        if st.button("🚀 PROCESAR PÓLIZAS", type="primary", use_container_width=True):

            # Contenedores de progreso
            progress_container = st.container()
            status_container = st.container()
            result_container = st.container()

            with progress_container:
                progress_general = st.progress(0)
                status_text = st.empty()
                tiempo_text = st.empty()

            client = inicializar_cliente()
            resultados = {}
            tiempo_inicio = time.time()

            # Procesar cada PDF
            for idx_pdf, pdf_file in enumerate(uploaded_pdfs):
                try:
                    status_text.text(f"📄 Procesando {idx_pdf+1}/{len(uploaded_pdfs)}: {pdf_file.name}")

                    # Leer PDF
                    pdf_bytes = pdf_file.getvalue()

                    # Extraer texto según modo con el rango seleccionado
                    if modo_extraccion == "Texto nativo (PDF digital)":
                        texto = extraer_pdf_texto_completo(pdf_bytes, pagina_inicio, pagina_fin)
                    else:
                        texto = extraer_pdf_ocr_mejorado(pdf_bytes, pagina_inicio, pagina_fin, None)

                    if not texto or len(texto.strip()) < 100:
                        status_container.error(f"❌ No se pudo extraer texto suficiente de {pdf_file.name}")
                        continue

                    # Mostrar estadísticas
                    chars = len(texto)
                    paginas_detectadas = texto.count("--- PÁGINA")
                    status_text.text(f"🤖 Analizando {pdf_file.name} ({chars:,} caracteres, ~{paginas_detectadas} páginas)...")

                    # Procesar con IA por secciones
                    def update_status(msg):
                        status_text.text(f"🧠 {pdf_file.name}: {msg}")

                    elementos = estructurar_poliza_por_secciones(texto, pdf_file.name, client, update_status)

                    if elementos and len(elementos) > 0:
                        if validar_resultados:
                            validar_completitud(elementos, texto)

                        resultados[pdf_file.name] = elementos

                        with status_container:
                            st.markdown(f"""
                            <div class="success-box">
                                <strong>✅ {pdf_file.name}</strong><br>
                                • {len(elementos)} elementos extraídos<br>
                                • Tipos detectados: {', '.join(set(e.get('tipo','') for e in elementos))}
                            </div>
                            """, unsafe_allow_html=True)
                    else:
                        status_container.error(f"❌ No se estructuró contenido de {pdf_file.name}")

                    # Actualizar progreso
                    progress_general.progress((idx_pdf + 1) / len(uploaded_pdfs))

                    # Calcular tiempo restante
                    tiempo_transcurrido = time.time() - tiempo_inicio
                    if idx_pdf + 1 < len(uploaded_pdfs):
                        tiempo_por_archivo = tiempo_transcurrido / (idx_pdf + 1)
                        tiempo_restante = tiempo_por_archivo * (len(uploaded_pdfs) - (idx_pdf + 1))
                        tiempo_text.text(f"⏱️ Transcurrido: {tiempo_transcurrido/60:.1f}min | Estimado restante: {tiempo_restante/60:.1f}min")

                except Exception as e:
                    status_container.error(f"❌ Error procesando {pdf_file.name}: {str(e)}")
                    continue

            # Finalización
            progress_general.progress(1.0)
            status_text.text("✅ ¡Procesamiento completado!")

            if resultados:
                # Crear Excel
                excel_output = crear_excel_estructurado(resultados)

                with result_container:
                    st.subheader("📊 Resumen de Extracción")

                    # Métricas
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("PDFs procesados", len(resultados))
                    with col2:
                        total_elem = sum(len(d) for d in resultados.values())
                        st.metric("Elementos extraídos", total_elem)
                    with col3:
                        tiempo_total = time.time() - tiempo_inicio
                        st.metric("Tiempo total", f"{tiempo_total/60:.1f} min")
                    with col4:
                        avg_por_pdf = total_elem / len(resultados) if resultados else 0
                        st.metric("Promedio por PDF", f"{avg_por_pdf:.0f} elem")

                    # Vista previa
                    with st.expander("👁️ Vista previa del resultado", expanded=True):
                        primer_pdf = list(resultados.keys())[0]
                        st.write(f"**Primeras filas de: {primer_pdf}**")

                        df_preview = pd.DataFrame(resultados[primer_pdf])
                        if not df_preview.empty:
                            df_display = df_preview.head(10).copy()
                            st.dataframe(df_display, use_container_width=True)

                            if 'col_b' in df_preview.columns:
                                longitudes = df_preview['col_b'].str.len()
                                st.caption(f"Longitud de textos: min={longitudes.min()}, max={longitudes.max()}, avg={longitudes.mean():.0f} caracteres")

                    # Botón de descarga
                    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                    nombre_salida = f"polizas_estructuradas_{timestamp}.xlsx"

                    st.download_button(
                        label="📥 DESCARGAR EXCEL ESTRUCTURADO",
                        data=excel_output,
                        file_name=nombre_salida,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

                    st.balloons()
            else:
                result_container.error("❌ No se pudo procesar ningún PDF correctamente")

if __name__ == "__main__":
    main()
